# 提取模板管理 API（PDF 字段提取到 Excel）

import asyncio
import logging
from datetime import datetime
from io import BytesIO
from uuid import uuid4

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse

from app.pipeline.parsers.pdf_parser import DEFAULT_EXTRACTION_FIELDS
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_db, get_tenant
from app.models import Tenant, ExtractionSchema, KnowledgeBase, Document
from app.db.session import SessionLocal
from app.schemas.extraction_schema import (
    ExtractionSchemaResponse,
    ExtractionSchemaListResponse,
    BatchExtractResponse,
    ExtractedResult,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/v1/extraction-schemas", tags=["extraction"])


@router.post("", response_model=ExtractionSchemaResponse)
async def create_extraction_schema(
    file: UploadFile = File(..., description="Excel 模板文件（定义要提取的字段）"),
    name: str = Form(..., description="模板名称"),
    description: str | None = Form(default=None, description="模板描述"),
    knowledge_base_id: str | None = Form(default=None, description="关联的知识库 ID"),
    db: AsyncSession = Depends(get_db),
    tenant: Tenant = Depends(get_tenant),
):
    """
    创建提取模板

    上传 Excel 文件作为模板，系统会解析第一行作为字段名称。
    Excel 模板格式：
    - 第一行：字段名称（如：产品名称、价格、规格）
    - 第二行（可选）：字段类型（string/number/date/boolean）
    """
    # 验证文件类型
    filename = file.filename or "template.xlsx"
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext not in ("xlsx", "xls"):
        raise HTTPException(
            status_code=400, detail="仅支持 .xlsx 或 .xls 格式的 Excel 文件"
        )

    # 读取文件内容
    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="文件内容为空")

    # 解析 Excel 提取字段定义
    from app.pipeline.parsers.excel_parser import ExcelParser

    parser = ExcelParser()
    try:
        schema_obj = parser.extract_schema(file_bytes, name=name)
    except Exception as e:
        logger.error(f"解析 Excel 模板失败: {e}")
        raise HTTPException(status_code=400, detail=f"解析 Excel 模板失败: {str(e)}")

    if not schema_obj.fields:
        raise HTTPException(
            status_code=400,
            detail="未能从 Excel 中提取到字段定义，请确保第一行包含字段名称",
        )

    # 创建数据库记录
    extraction_schema = ExtractionSchema(
        id=str(uuid4()),
        tenant_id=tenant.id,
        knowledge_base_id=knowledge_base_id,
        name=name,
        description=description,
        fields=schema_obj.fields,
        source_filename=filename,
    )

    db.add(extraction_schema)
    await db.commit()
    await db.refresh(extraction_schema)

    logger.info(
        f"创建提取模板: {extraction_schema.id}, 字段数: {len(schema_obj.fields)}"
    )

    return extraction_schema


@router.get("", response_model=ExtractionSchemaListResponse)
async def list_extraction_schemas(
    knowledge_base_id: str | None = Query(default=None, description="按知识库过滤"),
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    tenant: Tenant = Depends(get_tenant),
):
    """列出提取模板"""
    query = select(ExtractionSchema).where(ExtractionSchema.tenant_id == tenant.id)

    if knowledge_base_id:
        query = query.where(ExtractionSchema.knowledge_base_id == knowledge_base_id)

    # 获取总数
    count_query = select(func.count()).select_from(query.subquery())
    total = (await db.execute(count_query)).scalar() or 0

    # 获取列表
    query = query.order_by(ExtractionSchema.created_at.desc()).offset(skip).limit(limit)
    result = await db.execute(query)
    items = result.scalars().all()

    # 转换为响应模型
    response_items = [
        ExtractionSchemaResponse(
            id=str(item.id),
            name=str(item.name),
            fields=item.fields or [],
            source_filename=str(item.source_filename) if item.source_filename else None,
            usage_count=str(item.usage_count) if item.usage_count else "0",
            created_at=item.created_at,
        )
        for item in items
    ]

    return ExtractionSchemaListResponse(items=response_items, total=total)


@router.get("/{schema_id}", response_model=ExtractionSchemaResponse)
async def get_extraction_schema(
    schema_id: str,
    db: AsyncSession = Depends(get_db),
    tenant: Tenant = Depends(get_tenant),
):
    """获取提取模板详情"""
    result = await db.execute(
        select(ExtractionSchema).where(
            ExtractionSchema.id == schema_id,
            ExtractionSchema.tenant_id == tenant.id,
        )
    )
    schema = result.scalar_one_or_none()

    if not schema:
        raise HTTPException(status_code=404, detail="提取模板不存在")

    # 转换为 Pydantic 模型
    return ExtractionSchemaResponse.model_validate(schema)


@router.delete("/{schema_id}")
async def delete_extraction_schema(
    schema_id: str,
    db: AsyncSession = Depends(get_db),
    tenant: Tenant = Depends(get_tenant),
):
    """删除提取模板"""
    result = await db.execute(
        select(ExtractionSchema).where(
            ExtractionSchema.id == schema_id,
            ExtractionSchema.tenant_id == tenant.id,
        )
    )
    schema = result.scalar_one_or_none()

    if not schema:
        raise HTTPException(status_code=404, detail="提取模板不存在")

    await db.delete(schema)
    await db.commit()

    return {"message": "删除成功"}


@router.post("/{schema_id}/extract")
async def extract_from_pdfs(
    schema_id: str,
    files: list[UploadFile] = File(..., description="PDF 文件列表"),
    output_format: str = Form(default="json", description="输出格式: json / excel"),
    kb_id: str | None = Form(
        default=None, description="知识库 ID（可选，指定后将 PDF 内容异步入库）"
    ),
    db: AsyncSession = Depends(get_db),
    tenant: Tenant = Depends(get_tenant),
):
    """
    批量从 PDF 文件中提取字段

    根据提取模板定义的字段，从上传的 PDF 文件中提取信息。
    支持两种输出格式：
    - json: 返回 JSON 格式的提取结果
    - excel: 返回 Excel 文件下载

    如果指定 kb_id，PDF 内容将异步存储到 OSS 并入库到知识库。
    """
    # 获取提取模板
    result = await db.execute(
        select(ExtractionSchema).where(
            ExtractionSchema.id == schema_id,
            ExtractionSchema.tenant_id == tenant.id,
        )
    )
    schema = result.scalar_one_or_none()

    if not schema:
        raise HTTPException(status_code=404, detail="提取模板不存在")

    if not files:
        raise HTTPException(status_code=400, detail="请上传至少一个 PDF 文件")

    # 如果指定了 kb_id，验证知识库存在且属于当前租户
    kb = None
    if kb_id:
        kb_result = await db.execute(
            select(KnowledgeBase).where(
                KnowledgeBase.id == kb_id,
                KnowledgeBase.tenant_id == tenant.id,
            )
        )
        kb = kb_result.scalar_one_or_none()
        if not kb:
            raise HTTPException(status_code=404, detail="知识库不存在或无权访问")

    # 导入解析器和 LLM
    from app.pipeline.parsers.pdf_parser import PDFParser
    from app.pipeline.parsers.base import ExtractionSchema as SchemaObj

    # 构建 Schema 对象 - 从 ORM 对象提取值
    schema_obj = SchemaObj(
        id=schema.id._value if hasattr(schema.id, "_value") else str(schema.id),
        name=schema.name._value if hasattr(schema.name, "_value") else str(schema.name),
        fields=list(schema.fields) if schema.fields else [],
        source_file=str(schema.source_filename) if schema.source_filename else "",
    )

    parser = PDFParser()
    results = []
    success_count = 0
    failed_count = 0

    # 用于存储需要入库的文件数据（如果指定了 kb_id）
    files_to_ingest: list[dict] = []

    for file in files:
        filename = file.filename or "unknown.pdf"

        # 验证文件类型
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if ext != "pdf":
            results.append(
                ExtractedResult(
                    filename=filename,
                    success=False,
                    error="仅支持 PDF 文件",
                )
            )
            failed_count += 1
            continue

        try:
            # 读取文件
            file_bytes = await file.read()
            if not file_bytes:
                results.append(
                    ExtractedResult(
                        filename=filename,
                        success=False,
                        error="文件内容为空",
                    )
                )
                failed_count += 1
                continue

            # 解析 PDF 并提取字段（传入 tenant_id 和 db_session 用于获取 LLM 配置）
            parse_result = await parser.parse(
                file_bytes=file_bytes,
                filename=filename,
                extraction_schema=schema_obj,
                tenant_id=tenant.id,
                db_session=db,
            )

            if parse_result.extracted_fields:
                results.append(
                    ExtractedResult(
                        filename=filename,
                        success=True,
                        fields=parse_result.extracted_fields,
                    )
                )
                success_count += 1

                # 如果指定了 kb_id，保存数据用于异步入库
                if kb:
                    files_to_ingest.append(
                        {
                            "filename": filename,
                            "file_bytes": file_bytes,
                            "markdown_content": parse_result.content,  # PDF 转换后的 Markdown
                            "extracted_fields": parse_result.extracted_fields,
                        }
                    )
            else:
                results.append(
                    ExtractedResult(
                        filename=filename,
                        success=False,
                        error="未能提取到字段",
                    )
                )
                failed_count += 1

        except Exception as e:
            logger.error(f"处理文件 {filename} 失败: {e}")
            results.append(
                ExtractedResult(
                    filename=filename,
                    success=False,
                    error=str(e),
                )
            )
            failed_count += 1

    # 更新使用次数
    schema.usage_count = str(int(schema.usage_count or "0") + len(files))

    # 如果有文件需要入库，创建 Document 记录并启动异步任务
    doc_ids = []
    if kb and files_to_ingest:
        for file_data in files_to_ingest:
            # 创建 Document 记录（状态为 processing）
            new_doc = Document(
                tenant_id=tenant.id,
                knowledge_base_id=kb_id,
                title=file_data["filename"],
                source=f"extraction:{schema.name}",
                raw_content=file_data["markdown_content"],
                extra_metadata={
                    "extraction_schema_id": schema.id,
                    "extraction_schema_name": schema.name,
                    "extracted_fields": file_data["extracted_fields"],
                },
                processing_status="processing",
                processing_log=f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 文档已创建，正在异步入库...\n",
            )
            db.add(new_doc)
            await db.flush()
            doc_ids.append(new_doc.id)

            # 启动后台任务执行实际入库
            asyncio.create_task(
                _background_extraction_ingest(
                    tenant_id=tenant.id,
                    kb_id=kb_id,
                    doc_id=new_doc.id,
                    filename=file_data["filename"],
                    file_bytes=file_data["file_bytes"],
                    markdown_content=file_data["markdown_content"],
                )
            )

        logger.info(f"已创建 {len(doc_ids)} 个文档记录，开始异步入库到知识库 {kb_id}")

    await db.commit()

    # 根据输出格式返回
    if output_format == "excel":
        # 生成 Excel 文件
        excel_bytes = _generate_excel(schema, results)

        return StreamingResponse(
            BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": 'attachment; filename="extraction_result.xlsx"'
            },
        )
    else:
        # 返回 JSON
        return BatchExtractResponse(
            results=results,
            total=len(files),
            success=success_count,
            failed=failed_count,
        )


def _generate_excel(schema: ExtractionSchema, results: list[ExtractedResult]) -> bytes:
    """将提取结果生成 Excel 文件"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "提取结果"

    # 表头样式
    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"
    )

    # 写入表头
    headers = ["文件名"] + [f["name"] for f in schema.fields] + ["状态", "错误信息"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 写入数据行
    for row_idx, result in enumerate(results, start=2):
        ws.cell(row=row_idx, column=1, value=result.filename)

        if result.success and result.fields:
            for col, field in enumerate(schema.fields, start=2):
                field_name = field["name"]
                value = result.fields.get(field_name)
                ws.cell(row=row_idx, column=col, value=value)
            ws.cell(row=row_idx, column=len(headers) - 1, value="成功")
        else:
            ws.cell(row=row_idx, column=len(headers) - 1, value="失败")
            ws.cell(row=row_idx, column=len(headers), value=result.error)

    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    # 导出为 bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


async def _background_extraction_ingest(
    tenant_id: str,
    kb_id: str,
    doc_id: str,
    filename: str,
    file_bytes: bytes,
    markdown_content: str,
):
    """
    后台任务：将 PDF 内容入库到知识库

    1. 存储原始 PDF 到 OSS
    2. 执行文档入库（chunking、embedding、写入向量库）
    """
    from app.services.ingestion import ingest_document
    from app.schemas.internal import IngestionParams

    # 让出控制权，确保任务在正确的事件循环上下文中运行
    await asyncio.sleep(0)

    logger.info(f"[后台入库] 开始处理提取文档 {doc_id}: {filename}")

    async with SessionLocal() as db:
        try:
            # 获取知识库
            kb_result = await db.execute(
                select(KnowledgeBase).where(KnowledgeBase.id == kb_id)
            )
            kb = kb_result.scalar_one_or_none()
            if not kb:
                logger.error(f"[后台入库] 知识库 {kb_id} 不存在")
                return

            # 获取文档记录
            doc_result = await db.execute(select(Document).where(Document.id == doc_id))
            doc = doc_result.scalar_one_or_none()
            if not doc:
                logger.error(f"[后台入库] 文档 {doc_id} 不存在")
                return

            # 更新处理日志
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            doc.processing_log += f"[{ts}] [INFO] 开始入库处理\n"
            await db.flush()

            # 存储原始 PDF 到 OSS
            from app.services.file_storage import get_file_storage

            file_storage = get_file_storage()
            if file_storage.enabled:
                try:
                    raw_file_path = await file_storage.store_raw_file(
                        tenant_id=tenant_id,
                        doc_id=doc_id,
                        filename=filename,
                        content=file_bytes,
                    )
                    if raw_file_path:
                        doc.raw_file_path = raw_file_path
                        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        doc.processing_log += (
                            f"[{ts}] [INFO] 原始文件已存储到 OSS: {raw_file_path}\n"
                        )
                        logger.info(f"[后台入库] 原始文件已存储到 OSS: {raw_file_path}")
                except Exception as e:
                    logger.warning(f"[后台入库] 原始文件存储失败（不影响入库）: {e}")

            # 构建 IngestionParams
            params = IngestionParams(
                title=filename,
                content=markdown_content,
                metadata={
                    "source_type": "pdf_extraction",
                    "original_filename": filename,
                },
                source=f"extraction:{filename}",
                existing_doc_id=doc_id,
            )

            # 执行入库
            result = await ingest_document(
                db,
                tenant_id=tenant_id,
                kb=kb,
                params=params,
            )

            # 更新文档状态为完成
            doc.processing_status = "completed"
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            doc.processing_log += (
                f"[{ts}] [INFO] 入库完成，共 {len(result.chunks)} 个 chunks\n"
            )

            await db.commit()

            logger.info(
                f"[后台入库] 文档 {doc_id} 入库完成，共 {len(result.chunks)} 个 chunks"
            )

        except Exception as e:
            logger.error(f"[后台入库] 文档 {doc_id} 入库失败: {type(e).__name__}: {e}")
            try:
                # 更新文档状态为失败
                doc_result = await db.execute(
                    select(Document).where(Document.id == doc_id)
                )
                doc = doc_result.scalar_one_or_none()
                if doc:
                    doc.processing_status = "failed"
                    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    doc.processing_log += (
                        f"[{ts}] [ERROR] 入库失败: {type(e).__name__}: {e}\n"
                    )
                    await db.commit()
            except Exception as inner_e:
                logger.error(f"[后台入库] 更新文档状态失败: {inner_e}")


# =============================================================================
# Qwen-PLUS 直接提取接口（使用默认字段模板）
# =============================================================================


@router.post("/extract/qwen-plus")
async def extract_with_qwen_plus(
    file: UploadFile = File(..., description="PDF 文件"),
    db: AsyncSession = Depends(get_db),
    tenant: Tenant = Depends(get_tenant),
):
    """
    使用 Qwen3.5-PLUS 模型直接提取 PDF 字段

    使用默认的30字段模板，无需创建提取模板。
    使用阿里云 DashScope 的 qwen3.5-vl-plus 视觉模型（将PDF转为图片）。

    返回提取的结构化 JSON 数据。
    """
    from app.config import get_settings
    import base64

    filename = file.filename or "unknown.pdf"
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext != "pdf":
        raise HTTPException(status_code=400, detail="仅支持 PDF 文件")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="文件内容为空")

    settings = get_settings()
    qwen_api_key = settings.qwen_api_key
    qwen_base_url = (
        settings.qwen_api_base or "https://dashscope.aliyuncs.com/compatible-mode/v1"
    )

    if not qwen_api_key:
        raise HTTPException(status_code=500, detail="未配置 Qwen API Key")

    fields = DEFAULT_EXTRACTION_FIELDS
    field_list = "\n".join([f"- {f['name']}" for f in fields])

    prompt = f"""请从以下PDF文档的页面图片中提取所有项目的完整信息。

## 需要提取的字段（每个项目都需要提取以下所有字段）
{field_list}

## 重要说明
1. 一个PDF文档可能包含多个项目（如ND-003、ND-006、XY-001等），请提取**所有项目**
2. "项目"字段是项目编号/名称，如ND-003、ND-006等
3. "研发机构"是指项目的发起公司/研究机构
4. 如果某个字段在某个项目中未提及，填写"未提及"
5. 必须以JSON数组格式返回，每个元素是一个项目的完整信息

## 输出要求
1. 返回JSON数组格式，每个元素包含一个项目的所有字段
2. 例如: [{{"项目": "ND-003", "靶点": "...", "研发机构": "...", ...}}, {{"项目": "ND-006", ...}}]
3. 确保提取所有项目，不要遗漏
4. 每个项目的30个字段都必须填写

请直接返回JSON数组，不要包含任何其他文字："""

    try:
        import fitz
        import io

        doc = fitz.open(stream=file_bytes, filetype="pdf")
        page_count = len(doc)

        images_content = []
        for page_num in range(page_count):
            page = doc.load_page(page_num)
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
            img_bytes = pix.tobytes("png")
            img_b64 = base64.b64encode(img_bytes).decode("utf-8")
            images_content.append(
                {
                    "type": "image_url",
                    "image_url": {"url": f"data:image/png;base64,{img_b64}"},
                }
            )
        doc.close()

        if not images_content:
            raise HTTPException(status_code=400, detail="PDF 页面提取失败")

    except Exception as e:
        logger.error(f"PDF 转图片失败: {e}")
        raise HTTPException(status_code=500, detail=f"PDF 处理失败: {str(e)}")

    try:
        from openai import AsyncOpenAI

        client = AsyncOpenAI(api_key=qwen_api_key, base_url=qwen_base_url)

        messages = [
            {
                "role": "user",
                "content": [{"type": "text", "text": prompt}] + images_content,
            }
        ]

        response = await client.chat.completions.create(
            model="qwen3.5-vl-plus",
            messages=messages,
            temperature=0.1,
            max_tokens=8192,
        )

        content_response = response.choices[0].message.content
    except Exception as e:
        logger.error(f"Qwen3.5-PLUS 调用失败: {e}")
        raise HTTPException(status_code=500, detail=f"LLM 调用失败: {str(e)}")

    import json

    json_str = content_response.strip()

    if "```json" in json_str:
        json_str = json_str.split("```json")[1].split("```")[0]
    elif "```" in json_str:
        json_str = json_str.split("```")[1].split("```")[0]

    try:
        extracted_data = json.loads(json_str.strip())
    except json.JSONDecodeError as e:
        logger.warning(f"JSON 解析失败: {e}")
        extracted_data = {
            "_raw_response": content_response[:2000],
            "_parse_error": str(e),
        }

    return {
        "filename": filename,
        "page_count": page_count,
        "extracted_fields": extracted_data,
    }
