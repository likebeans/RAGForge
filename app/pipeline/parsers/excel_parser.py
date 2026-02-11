# Excel 文件解析器（xlsx/xls）

import logging
import uuid
from io import BytesIO
from .base import FileParser, ParseResult, ContentBlock, ContentType, ExtractionSchema

logger = logging.getLogger(__name__)


class ExcelParser(FileParser):
    """Excel 文件解析器（xlsx/xls）"""
    
    @property
    def supported_extensions(self) -> set[str]:
        return {".xlsx", ".xls"}
    
    async def parse(
        self,
        file_bytes: bytes,
        filename: str,
        extraction_schema: ExtractionSchema | None = None,
    ) -> ParseResult:
        """解析 Excel 文件"""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("请安装 openpyxl: uv add openpyxl")
        
        ext = self._get_extension(filename)
        
        # xls 格式需要 xlrd
        if ext == ".xls":
            return await self._parse_xls(file_bytes, filename)
        
        # xlsx 格式使用 openpyxl
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        
        blocks = []
        tables = []
        content_parts = []
        total_rows = 0
        
        for sheet in wb.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            
            # 过滤全空行
            rows = [r for r in rows if any(c is not None for c in r)]
            if not rows:
                continue
            
            total_rows += len(rows)
            
            # 提取表头（第一行）
            headers = [str(c) if c is not None else "" for c in rows[0]]
            # 去除尾部空列
            while headers and headers[-1] == "":
                headers.pop()
            
            if not headers:
                continue
            
            col_count = len(headers)
            
            # 构建 Markdown 表格
            md_lines = [f"## {sheet.title}\n"]
            md_lines.append("| " + " | ".join(headers) + " |")
            md_lines.append("|" + " --- |" * col_count)
            
            data_rows = []
            for row in rows[1:]:
                cells = [str(c) if c is not None else "" for c in row[:col_count]]
                # 补齐列数
                while len(cells) < col_count:
                    cells.append("")
                md_lines.append("| " + " | ".join(cells) + " |")
                data_rows.append(cells)
            
            md_content = "\n".join(md_lines)
            content_parts.append(md_content)
            
            # 保存结构化表格
            table_data = {
                "sheet": sheet.title,
                "headers": headers,
                "rows": data_rows,
                "row_count": len(data_rows),
            }
            tables.append(table_data)
            
            # 创建内容块
            blocks.append(ContentBlock(
                type=ContentType.TABLE,
                content=md_content,
                raw_data=table_data,
            ))
        
        return ParseResult(
            content="\n\n".join(content_parts),
            blocks=blocks,
            metadata={
                "format": "xlsx",
                "sheet_count": len(wb.worksheets),
                "table_count": len(tables),
                "total_rows": total_rows,
                "filename": filename,
            },
            tables=tables,
        )
    
    async def _parse_xls(self, file_bytes: bytes, filename: str) -> ParseResult:
        """解析旧版 xls 格式"""
        try:
            import xlrd
        except ImportError:
            raise ImportError("请安装 xlrd 以支持 .xls 格式: uv add xlrd")
        
        wb = xlrd.open_workbook(file_contents=file_bytes)
        
        blocks = []
        tables = []
        content_parts = []
        total_rows = 0
        
        for sheet in wb.sheets():
            if sheet.nrows == 0:
                continue
            
            total_rows += sheet.nrows
            
            # 提取表头
            headers = [str(sheet.cell_value(0, c)) for c in range(sheet.ncols)]
            while headers and headers[-1] == "":
                headers.pop()
            
            if not headers:
                continue
            
            col_count = len(headers)
            
            # 构建 Markdown 表格
            md_lines = [f"## {sheet.name}\n"]
            md_lines.append("| " + " | ".join(headers) + " |")
            md_lines.append("|" + " --- |" * col_count)
            
            data_rows = []
            for r in range(1, sheet.nrows):
                cells = [str(sheet.cell_value(r, c)) for c in range(col_count)]
                md_lines.append("| " + " | ".join(cells) + " |")
                data_rows.append(cells)
            
            md_content = "\n".join(md_lines)
            content_parts.append(md_content)
            
            table_data = {
                "sheet": sheet.name,
                "headers": headers,
                "rows": data_rows,
                "row_count": len(data_rows),
            }
            tables.append(table_data)
            
            blocks.append(ContentBlock(
                type=ContentType.TABLE,
                content=md_content,
                raw_data=table_data,
            ))
        
        return ParseResult(
            content="\n\n".join(content_parts),
            blocks=blocks,
            metadata={
                "format": "xls",
                "sheet_count": wb.nsheets,
                "table_count": len(tables),
                "total_rows": total_rows,
                "filename": filename,
            },
            tables=tables,
        )
    
    def extract_schema(self, file_bytes: bytes, name: str = "default") -> ExtractionSchema:
        """
        从 Excel 提取字段模板
        
        规则：
        - 第一行为字段名
        - 第二行如果是 string/number/date/boolean 则作为类型，否则忽略
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("请安装 openpyxl: uv add openpyxl")
        
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        
        fields = []
        seen_names = set()
        
        for sheet in wb.worksheets:
            rows = list(sheet.iter_rows(max_row=2, values_only=True))
            if not rows:
                continue
            
            headers = rows[0]
            types = rows[1] if len(rows) > 1 else [None] * len(headers)
            
            for i, header in enumerate(headers):
                if not header:
                    continue
                
                field_name = str(header).strip()
                if not field_name or field_name in seen_names:
                    continue
                
                seen_names.add(field_name)
                
                # 解析类型
                field_type = "string"
                if i < len(types) and types[i]:
                    type_str = str(types[i]).lower().strip()
                    if type_str in ("string", "text", "str", "文本"):
                        field_type = "string"
                    elif type_str in ("number", "int", "float", "数字", "整数", "小数"):
                        field_type = "number"
                    elif type_str in ("date", "datetime", "日期", "时间"):
                        field_type = "date"
                    elif type_str in ("boolean", "bool", "布尔", "是否"):
                        field_type = "boolean"
                
                fields.append({
                    "name": field_name,
                    "type": field_type,
                    "required": False,
                    "sheet": sheet.title,
                })
        
        return ExtractionSchema(
            id=str(uuid.uuid4()),
            name=name,
            fields=fields,
            source_file=self._current_filename if hasattr(self, '_current_filename') else "template.xlsx",
        )
