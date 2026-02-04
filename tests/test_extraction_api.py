"""测试 PDF 字段提取到 Excel API"""

import os
import requests
from io import BytesIO

API_BASE = os.getenv("API_BASE", "http://localhost:8020")
ADMIN_TOKEN = os.getenv("ADMIN_TOKEN", "admin-token-for-dev")


def create_test_tenant():
    """创建测试租户并获取 API Key"""
    # 先尝试获取已有租户
    resp = requests.get(
        f"{API_BASE}/admin/tenants",
        headers={"X-Admin-Token": ADMIN_TOKEN},
    )
    if resp.status_code == 200:
        tenants = resp.json().get("items", [])
        for t in tenants:
            if t["name"] == "extraction_test_tenant":
                # 创建新 API Key
                resp = requests.post(
                    f"{API_BASE}/admin/tenants/{t['id']}/api-keys",
                    headers={"X-Admin-Token": ADMIN_TOKEN},
                    json={"name": "test_key_extraction"},
                )
                if resp.status_code in (200, 201):
                    data = resp.json()
                    return data.get("api_key") or data.get("key")
    
    # 创建新租户
    resp = requests.post(
        f"{API_BASE}/admin/tenants",
        headers={"X-Admin-Token": ADMIN_TOKEN},
        json={"name": "extraction_test_tenant"},
    )
    if resp.status_code in (200, 201):
        data = resp.json()
        return data.get("initial_api_key") or data.get("api_key")
    
    print(f"创建租户失败: {resp.status_code} {resp.text}")
    return None


def create_test_excel_template():
    """创建测试用的 Excel 模板"""
    import openpyxl
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "提取字段"
    
    # 第一行：字段名
    ws["A1"] = "产品名称"
    ws["B1"] = "价格"
    ws["C1"] = "规格"
    ws["D1"] = "生产日期"
    
    # 第二行：字段类型
    ws["A2"] = "string"
    ws["B2"] = "number"
    ws["C2"] = "string"
    ws["D2"] = "date"
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def test_create_extraction_schema(api_key: str):
    """测试创建提取模板"""
    excel_bytes = create_test_excel_template()
    
    resp = requests.post(
        f"{API_BASE}/v1/extraction-schemas",
        headers={"Authorization": f"Bearer {api_key}"},
        files={"file": ("template.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"name": "产品信息提取模板", "description": "测试模板"},
    )
    
    print(f"创建提取模板: {resp.status_code}")
    if resp.status_code == 200:
        data = resp.json()
        print(f"  ID: {data['id']}")
        print(f"  字段数: {len(data['fields'])}")
        print(f"  字段: {data['fields']}")
        return data["id"]
    else:
        print(f"  错误: {resp.text}")
    return None


def test_list_extraction_schemas(api_key: str):
    """测试列出提取模板"""
    resp = requests.get(
        f"{API_BASE}/v1/extraction-schemas",
        headers={"Authorization": f"Bearer {api_key}"},
    )
    
    print(f"列出提取模板: {resp.status_code}")
    if resp.status_code == 200:
        data = resp.json()
        print(f"  总数: {data['total']}")
        for item in data["items"]:
            print(f"  - {item['name']} ({item['id']})")
    else:
        print(f"  错误: {resp.text}")


def test_get_extraction_schema(api_key: str, schema_id: str):
    """测试获取提取模板详情"""
    resp = requests.get(
        f"{API_BASE}/v1/extraction-schemas/{schema_id}",
        headers={"Authorization": f"Bearer {api_key}"},
    )
    
    print(f"获取提取模板: {resp.status_code}")
    if resp.status_code == 200:
        data = resp.json()
        print(f"  名称: {data['name']}")
        print(f"  字段: {data['fields']}")
    else:
        print(f"  错误: {resp.text}")


def create_test_pdf():
    """创建测试用的 PDF 文件（使用 reportlab）"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        
        # 添加产品信息
        c.setFont("Helvetica-Bold", 16)
        c.drawString(100, 750, "Product Information Sheet")
        
        c.setFont("Helvetica", 12)
        c.drawString(100, 700, "Product Name: iPhone 15 Pro Max")
        c.drawString(100, 680, "Price: 9999")
        c.drawString(100, 660, "Specifications: 256GB, Titanium Blue")
        c.drawString(100, 640, "Production Date: 2024-01-15")
        
        c.save()
        buffer.seek(0)
        return buffer.read()
    except ImportError:
        print("  [跳过] reportlab 未安装，无法创建测试 PDF")
        return None


def test_extract_from_pdf(api_key: str, schema_id: str):
    """测试从 PDF 提取字段"""
    pdf_bytes = create_test_pdf()
    if not pdf_bytes:
        return
    
    # JSON 格式输出
    print("  测试 JSON 输出...")
    resp = requests.post(
        f"{API_BASE}/v1/extraction-schemas/{schema_id}/extract",
        headers={"Authorization": f"Bearer {api_key}"},
        files=[("files", ("test_product.pdf", pdf_bytes, "application/pdf"))],
        data={"output_format": "json"},
    )
    
    print(f"  提取结果 (JSON): {resp.status_code}")
    if resp.status_code == 200:
        data = resp.json()
        print(f"    总数: {data['total']}, 成功: {data['success']}, 失败: {data['failed']}")
        for r in data["results"]:
            if r["success"]:
                print(f"    文件: {r['filename']}")
                print(f"    字段: {r['fields']}")
            else:
                print(f"    文件: {r['filename']} - 失败: {r['error']}")
    else:
        print(f"    错误: {resp.text[:200]}")
    
    # Excel 格式输出
    print("  测试 Excel 输出...")
    resp = requests.post(
        f"{API_BASE}/v1/extraction-schemas/{schema_id}/extract",
        headers={"Authorization": f"Bearer {api_key}"},
        files=[("files", ("test_product.pdf", pdf_bytes, "application/pdf"))],
        data={"output_format": "excel"},
    )
    
    print(f"  提取结果 (Excel): {resp.status_code}")
    if resp.status_code == 200:
        content_type = resp.headers.get("content-type", "")
        print(f"    Content-Type: {content_type}")
        print(f"    文件大小: {len(resp.content)} bytes")
        
        # 验证 Excel 文件
        try:
            import openpyxl
            wb = openpyxl.load_workbook(BytesIO(resp.content))
            ws = wb.active
            print(f"    工作表: {ws.title}")
            print(f"    表头: {[cell.value for cell in ws[1]]}")
            if ws.max_row > 1:
                print(f"    第一行数据: {[cell.value for cell in ws[2]]}")
        except Exception as e:
            print(f"    验证 Excel 失败: {e}")
    else:
        print(f"    错误: {resp.text[:200]}")


def main():
    print("=" * 50)
    print("PDF 字段提取到 Excel API 测试")
    print("=" * 50)
    
    # 1. 创建测试租户
    print("\n[1] 创建测试租户...")
    api_key = create_test_tenant()
    if not api_key:
        print("无法获取 API Key，退出测试")
        return
    print(f"  API Key: {api_key[:20]}...")
    
    # 2. 创建提取模板
    print("\n[2] 创建提取模板...")
    schema_id = test_create_extraction_schema(api_key)
    
    # 3. 列出提取模板
    print("\n[3] 列出提取模板...")
    test_list_extraction_schemas(api_key)
    
    # 4. 获取提取模板详情
    if schema_id:
        print("\n[4] 获取提取模板详情...")
        test_get_extraction_schema(api_key, schema_id)
    
    # 5. 测试 PDF 提取
    if schema_id:
        print("\n[5] 测试 PDF 提取...")
        test_extract_from_pdf(api_key, schema_id)
    
    print("\n" + "=" * 50)
    print("测试完成！")
    print("=" * 50)


if __name__ == "__main__":
    main()
